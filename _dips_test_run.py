# -*- coding: utf-8 -*-
"""master.xlsx へテスト2件を入力し、DIPS登録用CSVを生成する使い捨てテスト。
- 入力先: E:\\中小企業新事業進出促進補助金\\master.xlsx（A〜K列に記入）
- CSV出力: E:\\中小企業新事業進出促進補助金\\DIPS出力_テスト（本番Dropboxには触れない）
"""
from __future__ import annotations

import copy
import datetime as dt
from pathlib import Path

from openpyxl import load_workbook

from dips import config, export_log, intake, pipeline

MASTER = Path(r"E:\中小企業新事業進出促進補助金\master.xlsx")
TEST_OUT = Path(r"E:\中小企業新事業進出促進補助金\DIPS出力_テスト")

# テスト2件（master入力イメージ）
PEOPLE = [
    {"koushu": "更新講習", "grade": "一等", "appno": "1234567890",
     "name": "試験 花子", "birth": dt.date(1992, 3, 15), "addr": "東京都千代田区1-1",
     "teishi": "無", "comp": dt.date(2026, 6, 18), "issue": dt.date(2026, 6, 18),
     "state": "新規", "phys": ""},
    {"koushu": "失効再交付講習", "grade": "二等", "appno": "0987654321",
     "name": "検証 次郎", "birth": dt.date(1988, 7, 20), "addr": "東京都港区2-2",
     "teishi": "無", "comp": dt.date(2026, 6, 18), "issue": dt.date(2026, 6, 18),
     "state": "新規", "phys": ""},
]

PREFIX = {"更新講習": "UC", "失効再交付講習": "EL"}
JOTAI = {"新規": "1：新規", "上書き修正": "2：上書き修正", "無効化": "3：無効化"}


def write_master():
    """master.xlsx の A〜K にテスト行を書き込む（数式列はそのまま＝Excelで計算）。"""
    wb = load_workbook(MASTER)
    ws = wb["修了者"]
    for i, p in enumerate(PEOPLE):
        r = 2 + i
        vals = [p["koushu"], p["grade"], p["appno"], p["name"], p["birth"],
                p["addr"], p["teishi"], p["comp"], p["issue"], p["state"], p["phys"]]
        for j, v in enumerate(vals, start=1):
            c = ws.cell(row=r, column=j, value=v)
            if j in (5, 8, 9):
                c.number_format = "yyyy/mm/dd"
            if j in (3, 11):
                c.number_format = "@"
    wb.save(MASTER)
    print(f"[master] {MASTER.name} に {len(PEOPLE)} 件入力（行2〜{1+len(PEOPLE)}）")


def make_csv():
    settings, mapping, code_map, _ = config.load_all()
    org = settings.get("organization", {})
    today = dt.date.today()

    tmp = copy.deepcopy(settings)
    tmp["data_dir"] = str(TEST_OUT)
    paths = config.Paths(tmp)
    paths.ensure_dirs()

    records = []
    for i, p in enumerate(PEOPLE, start=1):
        form = {
            "name": p["name"], "birth_date": p["birth"], "kubun": p["grade"],
            "applicant_no": p["appno"], "complete_date": p["comp"],
            "issue_date": p["issue"], "serial": i, "prefix": PREFIX[p["koushu"]],
            "koushi": "", "genteikaijo": "", "koufu": "有", "reissue_date": None,
            "teishi": p["teishi"], "jotai": JOTAI[p["state"]],
            "office_code": org.get("office_code_default", ""), "biko": "",
        }
        rec = intake.compute_row(
            form, org, holidays=[], months=settings["validity"]["months"],
            minus_days=settings["validity"]["minus_days"],
            link_due_days=settings["business_days"]["link_due_days"])
        records.append(rec)
        print(f"  - {p['name']}: 証明書番号={rec['修了証明書番号(自動)']} "
              f"有効期間満了日={rec['有効期間満了日(自動)']}")

    history = export_log.read_log(paths.export_log)
    prepared = pipeline.prepare_rows(records, org, code_map, mapping, tmp,
                                     history_rows=history, today=today)
    report = pipeline.generate(
        prepared, tmp, paths, mapping,
        date_str=today.strftime("%Y%m%d"),
        now_str=dt.datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
        write_log=True)

    print(f"\n[CSV] 生成 {report.total_written} 件 / {len(report.files)} ファイル")
    for fr in report.files:
        print(f"  ファイル: {fr.path}")
        print(f"  セルフテスト: {'✅ 全合格' if fr.selftest.passed else '❌ 不合格'}")
        for c in fr.selftest.checks:
            print(f"    - {c.name}: {'OK' if c.ok else 'NG'}")
    if report.excluded:
        print("  ⚠️ 除外:", [(p.apply_id, list(p.integrity_errors)) for p in report.excluded])
    return report


if __name__ == "__main__":
    print("=== masterテスト入力 ＋ DIPS CSV作成 ===")
    write_master()
    print("\n=== DIPS登録用CSV生成 ===")
    rep = make_csv()
    print("\n=== 完了 ===")
