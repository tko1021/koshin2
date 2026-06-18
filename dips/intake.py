"""手入力（Web）からの修了者レコード生成と master.xlsx 台帳への追記。

フォーム入力 → 各項目を自動計算（修了証明書番号・有効期間満了日・DIPS連携期限）し、
mapping.yaml が参照する列名キーの dict を返す。返した dict はそのまま
pipeline.prepare_rows() に渡せる（CSV発行）。また master.xlsx へ台帳として追記する。

修了証明書番号 = 講習区分(UC/EL) + 機関コード4桁 + 発行年月(西暦下2桁+月2桁) + 固有番号4桁
"""
from __future__ import annotations

import datetime as _dt
from pathlib import Path

from openpyxl import Workbook, load_workbook

from . import dates

# master「修了者マスタ」シート構造（ヘッダーは mapping.yaml の参照名と一致させる）
MASTER_SHEET = "修了者マスタ"
HEADER_ROW = 3
DATA_START = 4

# 列順（ヘッダー名は空白を含めない＝flatten後と同一になるようにする）
COLUMNS = [
    "管理No",
    "氏名",
    "技能証明申請者番号(10文字以内)",
    "資格区分",
    "講習区分",
    "停止処分者向け講習受講",
    "事務所コード",
    "講習修了日",
    "修了証明書発行日",
    "固有番号",
    "修了証明書番号(自動)",
    "申請ID(自動=証明書番号)",
    "有効期間満了日(自動)",
    "登録種別(状態フラグ)",
    "DIPS連携期限(自動)",
    "登録状況",
]

_KOUSHU_LABEL = {"UC": "UC：更新講習", "EL": "EL：失効再交付講習"}


def build_shumeisho_no(prefix: str, kikan_code: str, issue_date: _dt.date, serial) -> str:
    """修了証明書番号を組み立てる（UC + 機関コード + 発行年月YYMM + 固有番号4桁）。"""
    yy = issue_date.year % 100
    return f"{prefix}{kikan_code}{yy:02d}{issue_date.month:02d}{int(serial):04d}"


def compute_row(form: dict, organization: dict, holidays,
                months: int = 3, minus_days: int = 1, link_due_days: int = 5) -> dict:
    """フォーム入力 dict から、master列名キーの完全な1レコードを生成する。

    form のキー: name, applicant_no, kubun(一等/二等), complete_date(date),
                 issue_date(date), serial(int), teishi(無/有/空), jotai(登録種別文字列),
                 office_code(str), prefix(UC/EL), kanri_no(任意), status(任意)
    """
    issue = form["issue_date"]
    prefix = form.get("prefix", "UC")
    kikan_code = organization.get("kikan_code", "")
    office = form.get("office_code") or organization.get("office_code_default", "")

    shumeisho = build_shumeisho_no(prefix, kikan_code, issue, form["serial"])
    expiry = dates.expiry_date(issue, months, minus_days)
    due = dates.add_business_days(issue, link_due_days, holidays or [])

    return {
        "管理No": form.get("kanri_no"),
        "氏名": form["name"],
        "技能証明申請者番号(10文字以内)": form["applicant_no"],
        "資格区分": form["kubun"],
        "講習区分": _KOUSHU_LABEL.get(prefix, prefix),
        "停止処分者向け講習受講": form.get("teishi", ""),
        "事務所コード": office,
        "講習修了日": form["complete_date"],
        "修了証明書発行日": issue,
        "固有番号": f"{int(form['serial']):04d}",
        "修了証明書番号(自動)": shumeisho,
        "申請ID(自動=証明書番号)": shumeisho,   # 申請ID = 証明書番号
        "有効期間満了日(自動)": expiry,
        "登録種別(状態フラグ)": form.get("jotai", "1：新規"),
        "DIPS連携期限(自動)": due,
        "登録状況": form.get("status", "未登録"),
    }


def _ymd(value) -> str:
    d = dates.to_date(value)
    return dates.date_to_ymd(d) if d is not None else ""


def build_ledger_row(form: dict, organization: dict,
                     months: int = 3, minus_days: int = 1) -> dict:
    """修了証明書発行台帳（別添19準拠・12列）の1行を生成する。

    有効期限 = 修了日 + 3か月 - 1日（再交付でも延伸しない）。
    """
    issue = form["issue_date"]
    comp = form["complete_date"]
    prefix = form.get("prefix", "UC")
    kikan = organization.get("kikan_code", "")
    shumeisho = build_shumeisho_no(prefix, kikan, issue, form["serial"])
    expiry = dates.expiry_date(comp, months, minus_days)
    reissue = form.get("reissue_date")
    return {
        "修了証明書番号": shumeisho,
        "氏名": form["name"],
        "生年月日": _ymd(form.get("birth_date")),
        "区分": form["kubun"],
        "限定解除事項": form.get("genteikaijo", ""),
        "修了日": _ymd(comp),
        "有効期限": _ymd(expiry),
        "交付の有無": form.get("koufu", "有"),
        "再交付年月日": _ymd(reissue) if reissue else "",
        "担当講師": form.get("koushi", ""),
        "登録更新講習機関コード": kikan,
        "備考": form.get("biko", ""),
    }


# ---- master.xlsx 台帳 ----

def _ensure_workbook(path: Path, organization: dict) -> Workbook:
    if path.exists():
        return load_workbook(path)
    wb = Workbook()
    ws = wb.active
    ws.title = MASTER_SHEET
    ws.cell(row=1, column=1, value="修了者マスタ（DIPS更新修了者情報）")
    for j, name in enumerate(COLUMNS, start=1):
        ws.cell(row=HEADER_ROW, column=j, value=name)
    s = wb.create_sheet("設定")
    s["A2"], s["B2"] = "機関コード", organization.get("kikan_code", "")
    s["A3"], s["B3"] = "機関名", organization.get("kikan_name", "")
    s["A4"], s["B4"] = "事務所コード既定", organization.get("office_code_default", "")
    h = wb.create_sheet("祝日")
    h["A1"], h["B1"] = "日付", "名称"
    return wb


def _data_row_count(ws) -> int:
    n = 0
    for r in range(DATA_START, ws.max_row + 1):
        a = ws.cell(row=r, column=1).value   # 管理No
        b = ws.cell(row=r, column=2).value   # 氏名
        if (a is None or str(a).strip() == "") and (b is None or str(b).strip() == ""):
            continue
        n += 1
    return n


def _max_kanri_no(ws) -> int:
    mx = 0
    for r in range(DATA_START, ws.max_row + 1):
        v = ws.cell(row=r, column=1).value
        try:
            mx = max(mx, int(float(v)))
        except (TypeError, ValueError):
            continue
    return mx


def append_to_master(path: str | Path, rows: list[dict], organization: dict) -> int:
    """rows を master.xlsx「修了者マスタ」へ追記する。無ければ新規作成。管理Noは自動採番。"""
    p = Path(path)
    p.parent.mkdir(parents=True, exist_ok=True)
    wb = _ensure_workbook(p, organization)
    ws = wb[MASTER_SHEET]
    existing = _data_row_count(ws)
    next_no = _max_kanri_no(ws) + 1
    r = HEADER_ROW + 1 + existing
    for i, row in enumerate(rows):
        if not row.get("管理No"):
            row = {**row, "管理No": next_no + i}
        for j, name in enumerate(COLUMNS, start=1):
            ws.cell(row=r, column=j, value=row.get(name))
        r += 1
    wb.save(p)
    wb.close()
    return len(rows)


def read_holidays(path: str | Path) -> list[_dt.date]:
    """master.xlsx「祝日」シートの日付一覧（無ければ空）。連携期限計算に使用。"""
    p = Path(path)
    if not p.exists():
        return []
    try:
        wb = load_workbook(p, data_only=True)
        if "祝日" not in wb.sheetnames:
            wb.close()
            return []
        ws = wb["祝日"]
        out: list[_dt.date] = []
        for r in range(2, ws.max_row + 1):
            d = dates.to_date(ws.cell(row=r, column=1).value)
            if d is not None:
                out.append(d)
        wb.close()
        return out
    except Exception:
        return []
