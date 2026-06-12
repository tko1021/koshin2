"""master.xlsx 読込（openpyxl, data_only=True）。

- ヘッダーは flatten（改行・空白除去）して列を突合する。
- 番号系の欠落防止のため、値は原則そのまま保持し、出力時に str 化する。
"""
from __future__ import annotations

import datetime as _dt
import re
from pathlib import Path

from openpyxl import load_workbook

from . import dates


def flatten_header(s) -> str:
    """ヘッダー文字列から改行・空白（半角/全角）を全て除去した突合キーを返す。"""
    return re.sub(r"[\s　]+", "", "" if s is None else str(s))


class MasterData:
    def __init__(self, records, settings_values, holidays):
        self.records = records              # list[dict]: {flat_header: value, "_row": int}
        self.settings_values = settings_values  # {"kikan_code":..,"kikan_name":..,"office_code_default":..}
        self.holidays = holidays            # list[datetime.date]


def _cell_str(value) -> str:
    if value is None:
        return ""
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).strip()


def read_master(path: str | Path, mapping: dict) -> MasterData:
    wb = load_workbook(filename=str(path), data_only=True, read_only=False)
    try:
        records = _read_records(wb, mapping)
        settings_values = _read_settings_cells(wb, mapping)
        holidays = _read_holidays(wb, mapping)
    finally:
        wb.close()
    return MasterData(records, settings_values, holidays)


def _read_records(wb, mapping: dict) -> list[dict]:
    m = mapping["master"]
    ws = wb[m["sheet"]]
    header_row = m["header_row"]
    data_start = m["data_start_row"]

    # 列インデックス（1始まり）→ flatten ヘッダー
    col_to_header: dict[int, str] = {}
    for cell in ws[header_row]:
        flat = flatten_header(cell.value)
        if flat:
            col_to_header[cell.column] = flat

    # 行が「存在する」判定に使う列（管理No）
    kanri_key = flatten_header(mapping["master_aux"]["kanri_no"])

    records: list[dict] = []
    for row in ws.iter_rows(min_row=data_start):
        rec: dict = {"_row": row[0].row}
        any_value = False
        for cell in row:
            header = col_to_header.get(cell.column)
            if not header:
                continue
            rec[header] = cell.value
            if cell.value not in (None, ""):
                any_value = True
        # 管理No が空＝データ無しの行とみなして打ち切り
        if not any_value or _cell_str(rec.get(kanri_key, "")) == "":
            # 末尾の空行群に達したら終了
            if not any_value:
                break
            continue
        records.append(rec)
    return records


def _read_settings_cells(wb, mapping: dict) -> dict:
    s = mapping["settings_sheet"]
    ws = wb[s["sheet"]]
    out = {}
    for key, addr in s["cells"].items():
        out[key] = _cell_str(ws[addr].value)
    return out


def _read_holidays(wb, mapping: dict) -> list[_dt.date]:
    h = mapping["holiday_sheet"]
    ws = wb[h["sheet"]]
    date_col = h["date_col"]
    start = h["data_start_row"]
    result: list[_dt.date] = []
    for r in range(start, ws.max_row + 1):
        v = ws[f"{date_col}{r}"].value
        d = dates.to_date(v)
        if d is not None:
            result.append(d)
    return result
