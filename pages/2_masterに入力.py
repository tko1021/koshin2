# -*- coding: utf-8 -*-
"""master.xlsx（修了者管理マスタ）へ Web フォームから修了者を追記するページ。

- 入力先: settings.yaml の data_dir/master_file（＝共有フォルダの master.xlsx）の「修了者」シート
- 書き込むのは入力列 A〜K のみ。自動計算列 L〜T の数式は build_master.py が
  あらかじめ500行ぶん展開済みなので、Excel で開けば証明書番号・有効期限等が自動表示される。
- openpyxl は数式を計算しないため、画面プレビューの証明書番号・有効期限は Python で別計算する。
"""
from __future__ import annotations

import datetime as dt

import pandas as pd
import streamlit as st
from openpyxl import load_workbook

from dips import certificate, config, dates, intake

st.set_page_config(page_title="masterに入力", layout="wide")

settings, mapping, code_map, paths = config.load_all()
org = settings.get("organization", {})
kikan = org.get("kikan_code", "")
MASTER = paths.master                       # data_dir/master.xlsx（共有フォルダ）
SHEET = "修了者"
MONTHS = settings["validity"]["months"]
MINUS = settings["validity"]["minus_days"]
today = dt.date.today()

# 「修了者」シートの入力列（A〜K）＝build_master.py と一致
INPUT_COLS = ["講習区分", "等級区分", "技能証明申請者番号", "氏名", "生年月日",
              "住所", "停止処分者向け講習 受講有無", "修了日", "証明書発行日",
              "状態", "身体適性検査証明書番号"]
PREFIX = {"更新講習": "UC", "失効再交付講習": "EL"}
NUM_ROWS = 500
DATE_COLS = {5, 8, 9}      # 生年月日/修了日/発行日
TEXT_COLS = {3, 11}        # 申請者番号/身体適性番号（先頭ゼロ保持）

# 自動計算列 L〜T の数式テンプレート（既存行を超えて追記する場合に補完するため）
CALC_FORMULAS = {
    12: '=IF($I{r}="","",TEXT(I{r},"yy")&TEXT(I{r},"mm"))',
    13: '=IF($I{r}="","",TEXT(COUNTIF($L$2:L{r},L{r}),"0000"))',
    14: '=IF($I{r}="","",IF(A{r}="更新講習","UC","EL")&設定!$B$1&L{r}&M{r})',
    15: '=IF($I{r}="","",EDATE(I{r},3)-1)',
    16: '=IF($I{r}="","",WORKDAY(I{r},5,祝日!$A$2:$A$400))',
    17: '=IF($B{r}="","",IF(B{r}="一等","1","2"))',
    18: '=IF($G{r}="","",IF(G{r}="無","1","2"))',
    19: '=IF($J{r}="","",IF(J{r}="新規",1,IF(J{r}="上書き修正",2,3)))',
    20: '=IF($I{r}="","",IF(K{r}="","PA000000000000",K{r}))',
}


def _load_ws():
    if not MASTER.exists():
        return None, None
    wb = load_workbook(MASTER)
    if SHEET not in wb.sheetnames:
        return wb, None
    return wb, wb[SHEET]


def _read_entries(ws) -> list[dict]:
    """A〜K に値がある行を読み出し、証明書番号・有効期限を Python 計算して返す。"""
    rows = []
    serial_by_ym: dict[str, int] = {}
    r = 2
    while ws.cell(row=r, column=1).value or ws.cell(row=r, column=4).value:
        koushu = ws.cell(row=r, column=1).value or ""
        issue = dates.to_date(ws.cell(row=r, column=9).value)
        cert = expiry = ""
        if issue:
            ym = f"{issue.year % 100:02d}{issue.month:02d}"
            serial_by_ym[ym] = serial_by_ym.get(ym, 0) + 1
            cert = intake.build_shumeisho_no(
                PREFIX.get(koushu, "UC"), kikan, issue, serial_by_ym[ym])
            expiry = dates.date_to_ymd(dates.expiry_date(issue, MONTHS, MINUS))
        rows.append({
            "行": r, "講習区分": koushu, "氏名": ws.cell(row=r, column=4).value or "",
            "発行日": dates.date_to_ymd(issue) if issue else "",
            "証明書番号(自動)": cert, "有効期限(自動)": expiry,
        })
        r += 1
    return rows


def _next_free_row(ws) -> int:
    r = 2
    while ws.cell(row=r, column=1).value or ws.cell(row=r, column=4).value:
        r += 1
    return r


def _preview_cert(koushu, issue, existing) -> tuple[str, str]:
    """送信前プレビュー：同一発行年月の既存件数+1で固有番号を見積もる。"""
    if not issue:
        return "", ""
    ym = f"{issue.year % 100:02d}{issue.month:02d}"
    same = sum(1 for e in existing
               if e["発行日"][:7].replace("/", "")[2:] == ym) if existing else 0
    serial = same + 1
    cert = intake.build_shumeisho_no(PREFIX.get(koushu, "UC"), kikan, issue, serial)
    expiry = dates.date_to_ymd(dates.expiry_date(issue, MONTHS, MINUS))
    return cert, expiry


# ============================================================
# 画面
# ============================================================
st.title("📋 master.xlsx に修了者を入力")
st.caption(f"保存先: {MASTER}　|　機関コード: {kikan or '(未設定)'}")

wb, ws = _load_ws()
if ws is None:
    st.error(f"master.xlsx または「{SHEET}」シートが見つかりません: {MASTER}\n"
             "build_master.py で master.xlsx を生成してください。")
    st.stop()

existing = _read_entries(ws)

with st.form("master_entry", clear_on_submit=True):
    st.subheader("修了者を入力")
    c1, c2, c3 = st.columns(3)
    koushu = c1.selectbox("講習区分 *", ["更新講習", "失効再交付講習"])
    grade = c2.selectbox("等級区分 *", ["一等", "二等"])
    appno = c3.text_input("技能証明申請者番号 *（10文字以内）", max_chars=10)

    c4, c5, c6 = st.columns(3)
    name = c4.text_input("氏名 *")
    birth = c5.date_input("生年月日", value=None, format="YYYY/MM/DD",
                          min_value=dt.date(1925, 1, 1), max_value=today)
    addr = c6.text_input("住所")

    c7, c8, c9 = st.columns(3)
    teishi = c7.selectbox("停止処分者向け講習 受講有無", ["無", "有"])
    comp = c8.date_input("修了日 *", value=today, format="YYYY/MM/DD")
    issue = c9.date_input("証明書発行日 *", value=today, format="YYYY/MM/DD")

    c10, c11 = st.columns(2)
    state = c10.selectbox("状態 *", ["新規", "上書き修正", "無効化"])
    phys = c11.text_input("身体適性検査証明書番号（空欄→PA000000000000）")

    st.markdown("**修了証明書用（任意）**")
    koushi = st.text_input("担当講師")
    st.caption("限定解除事項：区分ごとに選択（資格区分の列に記載されます）")
    g1, g2, g3 = st.columns(3)
    gentei = {
        "回転翼航空機（マルチ）": g1.multiselect("回転翼（マルチ）", certificate.GENTEI_ITEMS),
        "回転翼航空機（ヘリ）": g2.multiselect("回転翼（ヘリ）", certificate.GENTEI_ITEMS),
        "飛行機": g3.multiselect("飛行機", certificate.GENTEI_ITEMS),
    }

    pv_cert, pv_exp = _preview_cert(koushu, issue, existing)
    st.info(f"プレビュー　修了証明書番号: **{pv_cert}**　/　有効期限: **{pv_exp}**　"
            f"（固有番号は同月の登録順で自動採番）")

    submitted = st.form_submit_button("➕ master.xlsx に追記", type="primary")

if submitted:
    if not name or not appno:
        st.error("氏名と技能証明申請者番号は必須です。")
        st.stop()
    r = _next_free_row(ws)
    values = [koushu, grade, appno, name, birth, addr, teishi, comp, issue, state, phys]
    for j, v in enumerate(values, start=1):
        cell = ws.cell(row=r, column=j, value=(v if v not in (None, "") else None))
        if j in DATE_COLS and v:
            cell.number_format = "yyyy/mm/dd"
        if j in TEXT_COLS:
            cell.number_format = "@"
    # 担当講師(21)・限定解除（区分別 22〜24）
    ws.cell(row=r, column=21).value = koushi or None
    for air, col in {"回転翼航空機（マルチ）": 22, "回転翼航空機（ヘリ）": 23,
                     "飛行機": 24}.items():
        items = gentei.get(air) or []
        ws.cell(row=r, column=col).value = "、".join(items) if items else None
    # 既存の数式展開行(2〜501)を超える場合は数式も補完
    if r > 1 + NUM_ROWS:
        for col, tpl in CALC_FORMULAS.items():
            c = ws.cell(row=r, column=col, value=tpl.format(r=r))
            c.number_format = "yyyy/mm/dd" if col in (15, 16) else "@"
    try:
        wb.save(MASTER)
    except PermissionError:
        st.error("master.xlsx を保存できませんでした。Excel で開いたままになっていないか"
                 "確認して閉じてから、もう一度お試しください。")
        st.stop()
    st.success(f"追記しました（{r}行目）: {name}　{pv_cert}")
    st.rerun()

st.divider()
st.subheader(f"現在の登録（{len(existing)} 件）")
if existing:
    st.dataframe(pd.DataFrame(existing), hide_index=True, use_container_width=True)
else:
    st.write("まだ登録がありません。上のフォームから入力してください。")
