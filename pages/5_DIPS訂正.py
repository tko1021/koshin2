# -*- coding: utf-8 -*-
"""DIPS 登録済みの修了者を訂正（上書き修正/無効化）するページ。

対象は「すでに DIPS CSV を発行済み（履歴 export_log に申請IDがある）」の修了者。
- 上書き修正(状態フラグ2): 証明書番号（=申請ID）を変えない項目だけ修正し、同じ申請IDで再発行。
- 無効化(状態フラグ3): その登録を取り消す。確認チェック必須。

証明書番号は「講習区分・発行日・発行年月内の固有番号」で決まるため、これらは変更不可
（変更すると別人の登録になる）。master の「状態」は新規のまま据え置き、訂正は本ページ
からの1回限りの操作として CSV を出力する（通常発行ページが訂正を繰り返さないため）。
"""
from __future__ import annotations

import datetime as dt

import pandas as pd
import streamlit as st
from openpyxl import load_workbook

from dips import config, dates, export_log, intake, pipeline

st.set_page_config(page_title="DIPS訂正", layout="wide")

settings, mapping, code_map, paths = config.load_all()
org = settings.get("organization", {})
kikan = org.get("kikan_code", "")
MASTER = paths.master
SHEET = "修了者"
today = dt.date.today()

PREFIX = {"更新講習": "UC", "失効再交付講習": "EL"}
GRADE_OPTS = ["一等", "二等"]
TEISHI_OPTS = ["無", "有"]
DATE_COLS = {5, 8, 9}
TEXT_COLS = {3, 11}


def read_issued(ws, issued_ids):
    """master を読み、証明書番号が履歴(issued_ids)にある＝発行済みの行だけ返す。"""
    out = []
    serial_by_ym: dict[str, int] = {}
    r = 2
    while ws.cell(row=r, column=1).value or ws.cell(row=r, column=4).value:
        issue = dates.to_date(ws.cell(row=r, column=9).value)
        if issue is None:
            r += 1
            continue
        koushu = ws.cell(row=r, column=1).value or "更新講習"
        ym = f"{issue.year % 100:02d}{issue.month:02d}"
        serial_by_ym[ym] = serial_by_ym.get(ym, 0) + 1
        serial = serial_by_ym[ym]
        cert = intake.build_shumeisho_no(PREFIX.get(koushu, "UC"), kikan, issue, serial)
        if cert in issued_ids:
            out.append({
                "行": r, "koushu": koushu, "serial": serial, "cert": cert,
                "grade": ws.cell(row=r, column=2).value or "一等",
                "appno": ws.cell(row=r, column=3).value or "",
                "name": ws.cell(row=r, column=4).value or "",
                "teishi": ws.cell(row=r, column=7).value or "無",
                "comp": dates.to_date(ws.cell(row=r, column=8).value) or issue,
                "issue": issue,
                "phys": ws.cell(row=r, column=11).value or "",
            })
        r += 1
    return out


def _set(ws, r, col, value):
    cell = ws.cell(row=r, column=col)
    cell.value = value if value not in (None, "") else None
    if col in DATE_COLS and value:
        cell.number_format = "yyyy/mm/dd"
    if col in TEXT_COLS:
        cell.number_format = "@"


def issue_correction(rec, jotai_code, allow_invalidate):
    history = export_log.read_log(paths.export_log)
    prepared = pipeline.prepare_rows(
        [rec], org, code_map, mapping, settings,
        history_rows=history, today=today, allow_invalidate=allow_invalidate)
    report = pipeline.generate(
        prepared, settings, paths, mapping,
        date_str=today.strftime("%Y%m%d"),
        now_str=dt.datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
        write_log=True)
    return prepared, report


def show_report(prepared, report):
    if report.total_written:
        st.success(f"訂正CSVを発行しました: {report.total_written} 件 / {len(report.files)} ファイル")
    for fr in report.files:
        passed = "✅ 合格" if fr.selftest.passed else "❌ 不合格"
        st.markdown(f"**{fr.path}**（{fr.count}件）　セルフテスト: {passed}")
        st.table(pd.DataFrame([{"項目": c.name, "結果": "OK" if c.ok else "NG"}
                               for c in fr.selftest.checks]))
        try:
            with open(fr.path, "rb") as f:
                st.download_button("⬇️ このCSVをダウンロード", f.read(),
                                   file_name=fr.path.name, mime="text/csv")
        except OSError:
            pass
    if report.excluded:
        st.warning("発行できませんでした（検証エラー）:")
        for pr in report.excluded:
            reasons = list(pr.integrity_errors) + [pr.block_reason] if pr.blocked else list(pr.integrity_errors)
            st.write(f"- {pr.apply_id}: {' / '.join(r for r in reasons if r)}")


# ============================================================
# 画面
# ============================================================
st.title("🔧 DIPS登録済みの訂正（上書き修正・無効化）")
st.caption(f"対象: 発行済みの修了者　|　master: {MASTER}")

if not MASTER.exists():
    st.error(f"master.xlsx が見つかりません: {MASTER}")
    st.stop()
wb = load_workbook(MASTER)
if SHEET not in wb.sheetnames:
    st.error(f"「{SHEET}」シートがありません: {MASTER}")
    st.stop()

history = export_log.read_log(paths.export_log)
issued_ids = {str(h.get("申請ID") or "") for h in history if h.get("申請ID")}
issued = read_issued(wb[SHEET], issued_ids)

if not issued:
    st.info("発行済みの修了者がありません。先に「DIPS_CSV作成」ページでCSVを発行してください。")
    st.stop()

st.dataframe(pd.DataFrame([{"行": e["行"], "氏名": e["name"], "証明書番号(申請ID)": e["cert"],
                           "発行日": dates.date_to_ymd(e["issue"])} for e in issued]),
             hide_index=True, use_container_width=True)

labels = {f'{e["name"]}（{e["cert"]}）': i for i, e in enumerate(issued)}
sel = st.selectbox("訂正する修了者を選択", list(labels.keys()))
e = issued[labels[sel]]

st.info(f"🔒 変更不可（証明書番号を構成）：講習区分 **{e['koushu']}** ／ "
        f"発行日 **{dates.date_to_ymd(e['issue'])}** ／ 申請ID **{e['cert']}**")

mode = st.radio("訂正の種類", ["上書き修正（内容を直して再登録）", "無効化（登録を取り消す）"])


def build_rec(grade, appno, teishi, comp, phys, jotai):
    form = {
        "name": e["name"], "birth_date": None, "kubun": grade,
        "applicant_no": str(appno), "complete_date": comp, "issue_date": e["issue"],
        "serial": e["serial"], "prefix": PREFIX.get(e["koushu"], "UC"),
        "koushi": "", "genteikaijo": "", "koufu": "有", "reissue_date": None,
        "teishi": teishi, "jotai": jotai,
        "office_code": org.get("office_code_default", ""), "biko": "",
    }
    return intake.compute_row(
        form, org, holidays=[], months=settings["validity"]["months"],
        minus_days=settings["validity"]["minus_days"],
        link_due_days=settings["business_days"]["link_due_days"])


if mode.startswith("上書き修正"):
    with st.form("fix"):
        st.subheader("修正できる項目（証明書番号は変わりません）")
        c1, c2, c3 = st.columns(3)
        grade = c1.selectbox("等級区分", GRADE_OPTS, index=GRADE_OPTS.index(e["grade"]))
        appno = c2.text_input("技能証明申請者番号", value=e["appno"], max_chars=10)
        teishi = c3.selectbox("受講有無", TEISHI_OPTS, index=TEISHI_OPTS.index(e["teishi"]))
        c4, c5 = st.columns(2)
        comp = c4.date_input("修了日", value=e["comp"], format="YYYY/MM/DD")
        phys = c5.text_input("身体適性検査証明書番号", value=e["phys"])
        go = st.form_submit_button("🔧 上書き修正CSVを発行", type="primary")
    if go:
        if not appno:
            st.error("技能証明申請者番号は必須です。")
            st.stop()
        rec = build_rec(grade, appno, teishi, comp, phys, "2：上書き修正")
        if rec["修了証明書番号(自動)"] != e["cert"]:
            st.error(f"証明書番号が一致しません（想定 {e['cert']} / 生成 "
                     f"{rec['修了証明書番号(自動)']}）。master の並び順が変わった可能性があります。")
            st.stop()
        prepared, report = issue_correction(rec, "2", allow_invalidate=False)
        # master 側にも修正を反映（状態は新規のまま据え置き）
        if report.total_written:
            for col, val in ((2, grade), (3, appno), (7, teishi), (8, comp), (11, phys)):
                _set(wb[SHEET], e["行"], col, val)
            try:
                wb.save(MASTER)
            except PermissionError:
                st.warning("CSVは発行しましたが、master.xlsx へ修正を書き戻せませんでした"
                           "（Excelで開いたまま？）。後で「master編集」で揃えてください。")
        show_report(prepared, report)

else:  # 無効化
    st.warning("無効化すると、この修了者のDIPS登録が取り消されます。")
    confirm = st.checkbox(f"「{e['name']}（{e['cert']}）」を無効化することを確認しました")
    if st.button("🗑 無効化CSVを発行", type="primary", disabled=not confirm):
        rec = build_rec(e["grade"], e["appno"], e["teishi"], e["comp"], e["phys"], "3：無効化")
        prepared, report = issue_correction(rec, "3", allow_invalidate=True)
        show_report(prepared, report)
        if report.total_written:
            st.caption("※master の行はそのまま残ります（履歴の重複チェックにより通常発行で"
                       "再登録されることはありません）。行自体を消したい場合は「master編集」で削除してください。")
