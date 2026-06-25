# -*- coding: utf-8 -*-
"""master.xlsx の修了者について、更新講習 修了証明書 PDF を発行するページ。

- 1名ずつ、または全員分を一括でPDF生成
- 保存先: 共有フォルダ（data_dir）の「修了証明書」フォルダ
- 本番テンプレ画像（data_dir/修了証明書テンプレート.png）があれば自動でそれを背景に使用
"""
from __future__ import annotations

import datetime as dt
import io
import zipfile
from pathlib import Path

import pandas as pd
import streamlit as st
from openpyxl import load_workbook

from dips import certificate, config, dates, intake

st.set_page_config(page_title="修了証明書発行", layout="wide")

settings, mapping, code_map, paths = config.load_all()
org = settings.get("organization", {})
kikan = org.get("kikan_code", "")
kikan_name = org.get("kikan_name", "")
MASTER = paths.master
SHEET = "修了者"
MONTHS = settings["validity"]["months"]
MINUS = settings["validity"]["minus_days"]
CERT_DIR = Path(paths.data_dir) / "修了証明書"
PREFIX = {"更新講習": "UC", "失効再交付講習": "EL"}


def read_people(ws):
    out = []
    serial_by_ym: dict[str, int] = {}
    r = 2
    while ws.cell(row=r, column=1).value or ws.cell(row=r, column=4).value:
        issue = dates.to_date(ws.cell(row=r, column=9).value)
        if issue is None:
            r += 1
            continue
        koushu = ws.cell(row=r, column=1).value or "更新講習"
        ym = f"{issue.year % 100:02d}{issue.month:02d}"
        serial_by_ym[ym] = serial_by_ym.get(ym, 0) + 1
        cert_no = intake.build_shumeisho_no(PREFIX.get(koushu, "UC"), kikan, issue, serial_by_ym[ym])

        def _split(col):
            v = ws.cell(row=r, column=col).value
            return [s for s in str(v).replace(",", "、").split("、") if s.strip()] if v else []

        gentei_map = {
            "回転翼航空機（マルチ）": _split(22),
            "回転翼航空機（ヘリ）": _split(23),
            "飛行機": _split(24),
        }
        out.append({
            "cert_no": cert_no,
            "applicant_no": ws.cell(row=r, column=3).value or "",
            "name": ws.cell(row=r, column=4).value or "",
            "birth": dates.date_to_ymd(dates.to_date(ws.cell(row=r, column=5).value))
            if ws.cell(row=r, column=5).value else "",
            "koushu": koushu,
            "grade": ws.cell(row=r, column=2).value or "",
            "complete_date": dates.date_to_ymd(dates.to_date(ws.cell(row=r, column=8).value) or issue),
            "issue_date": dates.date_to_ymd(issue),
            "expiry": dates.date_to_ymd(dates.expiry_date(issue, MONTHS, MINUS)),
            "koushi": ws.cell(row=r, column=21).value or "",
            "gentei_map": gentei_map,
            "kikan_name": kikan_name,
            "address": org.get("address", ""),
            "kikan_code": kikan,
        })
        r += 1
    return out


def make_pdf(person, template, seal):
    safe_name = str(person["name"]).replace("　", "").replace(" ", "").replace("/", "_")
    out = CERT_DIR / f"{person['cert_no']}_{safe_name}.pdf"
    certificate.generate_certificate(person, out, template_image=template, seal_image=seal)
    return out


# ============================================================
# 画面
# ============================================================
st.title("📜 更新講習 修了証明書 発行（PDF）")
st.caption(f"保存先: {CERT_DIR}")

if not MASTER.exists():
    st.error(f"master.xlsx が見つかりません: {MASTER}")
    st.stop()
wb = load_workbook(MASTER)
if SHEET not in wb.sheetnames:
    st.error(f"「{SHEET}」シートがありません: {MASTER}")
    st.stop()

template = certificate.find_template(paths.data_dir)
seal = certificate.find_seal(paths.data_dir)
if template:
    st.success(f"本番テンプレート（背景）を使用します: {template.name}")
else:
    st.info("背景テンプレート未設置のため、正式レイアウトを描画して発行します。"
            f"　背景画像で差し替える場合は `{Path(paths.data_dir) / '修了証明書テンプレート.png'}` を置いてください。")
if seal:
    st.success(f"印影画像を使用します: {seal.name}")
else:
    st.info("印影画像が未設置のため、代表者欄に「印」枠を表示します。"
            f"　差し替えるには `{Path(paths.data_dir) / '印影.png'}` を置いてください（背景透過PNG推奨）。")

people = read_people(wb[SHEET])
if not people:
    st.warning("master.xlsx に修了者がいません。「masterに入力」ページで登録してください。")
    st.stop()

def _gentei_summary(gmap, grade=""):
    # 表示ロジックはPDF（certificate.display_gentei）と共通。マルチのみ・「・」で1行連結。
    return "・".join(certificate.display_gentei(gmap, grade))


st.dataframe(pd.DataFrame([{"証明書番号": p["cert_no"], "氏名": p["name"], "資格区分": p["grade"],
                           "修了日": p["complete_date"], "有効期限": p["expiry"],
                           "担当講師": p["koushi"], "限定解除": _gentei_summary(p["gentei_map"], p["grade"])}
                          for p in people]),
             hide_index=True, use_container_width=True)
st.caption("※担当講師・限定解除事項は「masterに入力」「master編集」ページで人ごとに登録した内容を使用します。")

st.divider()
labels = ["▼ 全員分を一括発行"] + [f'{p["name"]}（{p["cert_no"]}）' for p in people]
sel = st.selectbox("発行する対象", labels)

if st.button("📜 修了証明書を発行（PDF）", type="primary"):
    if sel == "▼ 全員分を一括発行":
        targets = people
    else:
        targets = [people[labels.index(sel) - 1]]

    made = []
    for p in targets:
        try:
            made.append(make_pdf(p, template, seal))
        except Exception as e:  # noqa: BLE001
            st.error(f"{p['name']} の発行に失敗: {e}")
    if not made:
        st.stop()

    st.success(f"{len(made)} 件の修了証明書PDFを発行しました（保存先: {CERT_DIR}）")

    if len(made) == 1:
        with open(made[0], "rb") as f:
            st.download_button("⬇️ PDFをダウンロード", f.read(),
                               file_name=made[0].name, mime="application/pdf")
    else:
        buf = io.BytesIO()
        with zipfile.ZipFile(buf, "w", zipfile.ZIP_DEFLATED) as z:
            for p in made:
                z.write(p, arcname=p.name)
        st.download_button("⬇️ 全員分をZIPでダウンロード", buf.getvalue(),
                           file_name=f"修了証明書_{dt.date.today():%Y%m%d}.zip",
                           mime="application/zip")
        st.write("発行したファイル:")
        for p in made:
            st.write(f"- {p.name}")
