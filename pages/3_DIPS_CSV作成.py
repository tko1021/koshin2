# -*- coding: utf-8 -*-
"""master.xlsx（修了者シート）から DIPS 更新修了者情報の登録用CSV（公式17列）を発行するページ。

流れ: master.xlsx を読む → 各行を DIPS レコードへ変換 → koshin pipeline で
公式17列CSV（UTF-8 BOM / CRLF / 末尾改行なし）を生成＋セルフテスト。
すでに発行済みの申請ID（data/export_log.csv に記録）は自動で除外される。
"""
from __future__ import annotations

import datetime as dt

import pandas as pd
import streamlit as st
from openpyxl import load_workbook

from dips import config, dates, export_log, intake, pipeline

st.set_page_config(page_title="DIPS CSV作成", layout="wide")

settings, mapping, code_map, paths = config.load_all()
org = settings.get("organization", {})
MASTER = paths.master
SHEET = "修了者"
today = dt.date.today()

PREFIX = {"更新講習": "UC", "失効再交付講習": "EL"}
JOTAI = {"新規": "1：新規", "上書き修正": "2：上書き修正", "無効化": "3：無効化"}


def read_master_records(ws):
    """master「修了者」シートの入力行を DIPS レコード（compute_row出力）へ変換。

    固有番号は master の数式（同一発行年月で出現順）と一致させるため、行順で採番する。
    """
    recs, previews = [], []
    serial_by_ym: dict[str, int] = {}
    r = 2
    while ws.cell(row=r, column=1).value or ws.cell(row=r, column=4).value:
        koushu = ws.cell(row=r, column=1).value or "更新講習"
        grade = ws.cell(row=r, column=2).value or "一等"
        appno = ws.cell(row=r, column=3).value or ""
        name = ws.cell(row=r, column=4).value or ""
        teishi = ws.cell(row=r, column=7).value or "無"
        comp = dates.to_date(ws.cell(row=r, column=8).value)
        issue = dates.to_date(ws.cell(row=r, column=9).value)
        state = ws.cell(row=r, column=10).value or "新規"
        phys = ws.cell(row=r, column=11).value or ""
        if issue is None:
            r += 1
            continue
        ym = f"{issue.year % 100:02d}{issue.month:02d}"
        serial_by_ym[ym] = serial_by_ym.get(ym, 0) + 1
        form = {
            "name": name, "birth_date": None, "kubun": grade,
            "applicant_no": str(appno), "complete_date": comp or issue,
            "issue_date": issue, "serial": serial_by_ym[ym],
            "prefix": PREFIX.get(koushu, "UC"), "koushi": "", "genteikaijo": "",
            "koufu": "有", "reissue_date": None, "teishi": teishi,
            "jotai": JOTAI.get(state, "1：新規"),
            "office_code": org.get("office_code_default", ""), "biko": "",
        }
        rec = intake.compute_row(
            form, org, holidays=[], months=settings["validity"]["months"],
            minus_days=settings["validity"]["minus_days"],
            link_due_days=settings["business_days"]["link_due_days"])
        recs.append(rec)
        previews.append({
            "行": r, "氏名": name, "講習区分": koushu,
            "証明書番号": rec["修了証明書番号(自動)"],
            "発行日": dates.date_to_ymd(issue),
            "有効期間満了日": dates.date_to_ymd(rec["有効期間満了日(自動)"]),
            "状態": state,
        })
        r += 1
    return recs, previews


# ============================================================
# 画面
# ============================================================
st.title("📤 DIPS登録用CSVを作成")
st.caption(f"入力元: {MASTER}　|　CSV出力先: {paths.output_dir}　|　履歴: {paths.export_log}")

if not MASTER.exists():
    st.error(f"master.xlsx が見つかりません: {MASTER}")
    st.stop()
wb = load_workbook(MASTER)
if SHEET not in wb.sheetnames:
    st.error(f"「{SHEET}」シートがありません: {MASTER}")
    st.stop()

records, previews = read_master_records(wb[SHEET])

st.subheader(f"master の登録内容（{len(previews)} 件）")
if not previews:
    st.info("master.xlsx に修了者がありません。先に「masterに入力」ページで登録してください。")
    st.stop()
st.dataframe(pd.DataFrame(previews), hide_index=True, use_container_width=True)

# 発行済み（履歴にある申請ID）を表示
history = export_log.read_log(paths.export_log)
issued_ids = {str(h.get("申請ID") or h.get("apply_id") or "") for h in history}
new_recs = [r for r in records if r["申請ID(自動=証明書番号)"] not in issued_ids]
st.write(f"うち **未発行: {len(new_recs)} 件**　／　発行済み(履歴で自動除外): "
         f"{len(records) - len(new_recs)} 件")

st.divider()
st.markdown("**CSVを発行すると、未発行の申請のみが公式17列CSVとして出力され、"
            "履歴に記録されます（次回以降は自動で除外）。**")

if st.button("📤 DIPS登録用CSV を発行", type="primary", disabled=not new_recs):
    prepared = pipeline.prepare_rows(
        records, org, code_map, mapping, settings,
        history_rows=history, today=today)
    report = pipeline.generate(
        prepared, settings, paths, mapping,
        date_str=today.strftime("%Y%m%d"),
        now_str=dt.datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
        write_log=True)

    st.success(f"CSV生成: {report.total_written} 件 / {len(report.files)} ファイル")
    for fr in report.files:
        passed = "✅ 合格" if fr.selftest.passed else "❌ 不合格"
        st.markdown(f"**{fr.path}**（{fr.count}件）　セルフテスト: {passed}")
        st.table(pd.DataFrame([{"項目": c.name, "結果": "OK" if c.ok else "NG",
                                "詳細": c.detail} for c in fr.selftest.checks]))
        try:
            with open(fr.path, "rb") as f:
                st.download_button("⬇️ このCSVをダウンロード", f.read(),
                                   file_name=fr.path.name, mime="text/csv")
        except OSError:
            pass

    if report.excluded:
        st.warning("以下はDIPS検証エラー等で除外されました:")
        rows_err = []
        for pr in report.excluded:
            reasons = list(pr.integrity_errors)
            for e in pr.field_errors:
                reasons.append(f"[列{e.index} {e.name}] {e.msg}")
            rows_err.append({"申請ID": pr.apply_id, "理由": " / ".join(reasons)})
        st.table(pd.DataFrame(rows_err))
