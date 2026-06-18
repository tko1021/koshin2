# -*- coding: utf-8 -*-
"""master.xlsx（修了者シート）の既存行を編集・削除するページ（入力ミス訂正用）。

⚠️ 固有番号は「同一発行年月の入力順」で自動採番されるため、行の削除や発行日の
月変更を行うと同月の他者の証明書番号がずれる。よって本ページは「まだ DIPS CSV を
発行していない入力ミスの訂正」専用とする。発行済みの訂正は状態フラグ（上書き修正/
無効化）で別途行うこと。
"""
from __future__ import annotations

import datetime as dt

import pandas as pd
import streamlit as st
from openpyxl import load_workbook

from dips import certificate, config, dates, intake

GENTEI_COLS = {"回転翼航空機（マルチ）": 22, "回転翼航空機（ヘリ）": 23, "飛行機": 24}


def _split_items(v):
    return [s for s in str(v).replace(",", "、").split("、") if s.strip()] if v else []

st.set_page_config(page_title="master編集", layout="wide")

settings, mapping, code_map, paths = config.load_all()
org = settings.get("organization", {})
kikan = org.get("kikan_code", "")
MASTER = paths.master
SHEET = "修了者"
MONTHS = settings["validity"]["months"]
MINUS = settings["validity"]["minus_days"]
today = dt.date.today()

PREFIX = {"更新講習": "UC", "失効再交付講習": "EL"}
KOUSHU_OPTS = ["更新講習", "失効再交付講習"]
GRADE_OPTS = ["一等", "二等"]
TEISHI_OPTS = ["無", "有"]
STATE_OPTS = ["新規", "上書き修正", "無効化"]
DATE_COLS = {5, 8, 9}
TEXT_COLS = {3, 11}


def read_all(ws):
    """A〜K に値がある行を全フィールド読み出す。"""
    out = []
    serial_by_ym: dict[str, int] = {}
    r = 2
    while ws.cell(row=r, column=1).value or ws.cell(row=r, column=4).value:
        issue = dates.to_date(ws.cell(row=r, column=9).value)
        cert = ""
        if issue:
            ym = f"{issue.year % 100:02d}{issue.month:02d}"
            serial_by_ym[ym] = serial_by_ym.get(ym, 0) + 1
            cert = intake.build_shumeisho_no(
                PREFIX.get(ws.cell(row=r, column=1).value or "更新講習", "UC"),
                kikan, issue, serial_by_ym[ym])
        out.append({
            "行": r,
            "koushu": ws.cell(row=r, column=1).value or "更新講習",
            "grade": ws.cell(row=r, column=2).value or "一等",
            "appno": ws.cell(row=r, column=3).value or "",
            "name": ws.cell(row=r, column=4).value or "",
            "birth": dates.to_date(ws.cell(row=r, column=5).value),
            "addr": ws.cell(row=r, column=6).value or "",
            "teishi": ws.cell(row=r, column=7).value or "無",
            "comp": dates.to_date(ws.cell(row=r, column=8).value),
            "issue": issue,
            "state": ws.cell(row=r, column=10).value or "新規",
            "phys": ws.cell(row=r, column=11).value or "",
            "koushi": ws.cell(row=r, column=21).value or "",
            "gentei": {air: _split_items(ws.cell(row=r, column=col).value)
                       for air, col in GENTEI_COLS.items()},
            "cert": cert,
        })
        r += 1
    return out


def write_extra(ws, r, koushi, gentei):
    """担当講師(21)・限定解除(22〜24) を書く。gentei={機体:[項目]}。"""
    ws.cell(row=r, column=21).value = koushi or None
    for air, col in GENTEI_COLS.items():
        items = (gentei or {}).get(air) or []
        ws.cell(row=r, column=col).value = "、".join(items) if items else None


def write_row(ws, r, vals):
    # 注意: ws.cell(..., value=None) は openpyxl では「何もしない」ため、
    # 空欄へ戻す/クリアするには cell.value = None と明示代入する必要がある。
    for j, v in enumerate(vals, start=1):
        cell = ws.cell(row=r, column=j)
        cell.value = v if v not in (None, "") else None
        if j in DATE_COLS and v:
            cell.number_format = "yyyy/mm/dd"
        if j in TEXT_COLS:
            cell.number_format = "@"


def rewrite_all(ws, entries):
    """entries（dictのlist）を行2から書き直し、余った末尾行の入力列を消す。"""
    last_used = 1
    for i, e in enumerate(entries):
        r = 2 + i
        write_row(ws, r, [e["koushu"], e["grade"], e["appno"], e["name"], e["birth"],
                          e["addr"], e["teishi"], e["comp"], e["issue"], e["state"],
                          e["phys"]])
        write_extra(ws, r, e.get("koushi", ""), e.get("gentei"))
        last_used = r
    # 末尾の余り行（旧データ）の入力列A〜K・21〜24をクリア（数式列L〜Tは残す）
    r = last_used + 1
    while ws.cell(row=r, column=1).value or ws.cell(row=r, column=4).value:
        for j in list(range(1, 12)) + [21, 22, 23, 24]:
            ws.cell(row=r, column=j).value = None   # 明示代入でクリア
        r += 1


# ============================================================
# 画面
# ============================================================
st.title("✏️ master を編集・削除（入力ミス訂正）")
st.caption(f"対象: {MASTER}")
st.warning("⚠️ このページは **DIPS CSV を発行する前**の入力ミス訂正用です。"
           "削除や発行日の月変更は同月の他者の証明書番号をずらす可能性があります。"
           "発行済みの訂正は状態フラグ（上書き修正/無効化）で行ってください。")

if not MASTER.exists():
    st.error(f"master.xlsx が見つかりません: {MASTER}")
    st.stop()
wb = load_workbook(MASTER)
if SHEET not in wb.sheetnames:
    st.error(f"「{SHEET}」シートがありません: {MASTER}")
    st.stop()

entries = read_all(wb[SHEET])
if not entries:
    st.info("master.xlsx に修了者がありません。「masterに入力」ページで登録してください。")
    st.stop()

st.dataframe(pd.DataFrame([{"行": e["行"], "氏名": e["name"], "講習区分": e["koushu"],
                           "証明書番号": e["cert"],
                           "発行日": dates.date_to_ymd(e["issue"]) if e["issue"] else "",
                           "状態": e["state"]} for e in entries]),
             hide_index=True, use_container_width=True)

# 編集対象の選択
labels = {f'{e["行"]}行: {e["name"]}（{e["cert"]}）': i for i, e in enumerate(entries)}
sel = st.selectbox("編集・削除する修了者を選択", list(labels.keys()))
idx = labels[sel]
e = entries[idx]

st.divider()
with st.form("edit"):
    st.subheader(f"{e['行']}行目を編集")
    c1, c2, c3 = st.columns(3)
    koushu = c1.selectbox("講習区分", KOUSHU_OPTS, index=KOUSHU_OPTS.index(e["koushu"]))
    grade = c2.selectbox("等級区分", GRADE_OPTS, index=GRADE_OPTS.index(e["grade"]))
    appno = c3.text_input("技能証明申請者番号", value=e["appno"], max_chars=10)

    c4, c5, c6 = st.columns(3)
    name = c4.text_input("氏名", value=e["name"])
    birth = c5.date_input("生年月日", value=e["birth"], format="YYYY/MM/DD",
                          min_value=dt.date(1925, 1, 1), max_value=today)
    addr = c6.text_input("住所", value=e["addr"])

    c7, c8, c9 = st.columns(3)
    teishi = c7.selectbox("受講有無", TEISHI_OPTS, index=TEISHI_OPTS.index(e["teishi"]))
    comp = c8.date_input("修了日", value=e["comp"] or today, format="YYYY/MM/DD")
    issue = c9.date_input("証明書発行日", value=e["issue"] or today, format="YYYY/MM/DD")

    c10, c11 = st.columns(2)
    state = c10.selectbox("状態", STATE_OPTS, index=STATE_OPTS.index(e["state"]))
    phys = c11.text_input("身体適性検査証明書番号", value=e["phys"])

    st.markdown("**修了証明書用**")
    koushi = st.text_input("担当講師", value=e.get("koushi", ""))
    g1, g2, g3 = st.columns(3)
    gentei = {
        "回転翼航空機（マルチ）": g1.multiselect(
            "回転翼（マルチ）", certificate.GENTEI_ITEMS,
            default=e["gentei"].get("回転翼航空機（マルチ）", [])),
        "回転翼航空機（ヘリ）": g2.multiselect(
            "回転翼（ヘリ）", certificate.GENTEI_ITEMS,
            default=e["gentei"].get("回転翼航空機（ヘリ）", [])),
        "飛行機": g3.multiselect(
            "飛行機", certificate.GENTEI_ITEMS, default=e["gentei"].get("飛行機", [])),
    }

    update = st.form_submit_button("💾 この内容で更新", type="primary")

if update:
    if not name or not appno:
        st.error("氏名と技能証明申請者番号は必須です。")
        st.stop()
    write_row(wb[SHEET], e["行"],
              [koushu, grade, appno, name, birth, addr, teishi, comp, issue, state, phys])
    write_extra(wb[SHEET], e["行"], koushi, gentei)
    try:
        wb.save(MASTER)
    except PermissionError:
        st.error("保存できませんでした。master.xlsx を Excel で閉じてから再実行してください。")
        st.stop()
    st.success(f"{e['行']}行目を更新しました: {name}")
    st.rerun()

st.divider()
st.subheader("🗑 削除")
st.write(f"選択中: **{e['行']}行 / {e['name']}（{e['cert']}）**")
confirm = st.checkbox("この行を削除することを確認しました")
if st.button("🗑 削除する", disabled=not confirm):
    remaining = [x for i, x in enumerate(entries) if i != idx]
    rewrite_all(wb[SHEET], remaining)
    try:
        wb.save(MASTER)
    except PermissionError:
        st.error("保存できませんでした。master.xlsx を Excel で閉じてから再実行してください。")
        st.stop()
    st.success(f"{e['行']}行目（{e['name']}）を削除しました。")
    st.rerun()
